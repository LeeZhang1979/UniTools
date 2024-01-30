#!/usr/bin/env python3
# coding=utf-8

'''线束表转化'''

from asyncio.format_helpers import _format_args_and_kwargs
import os
from pyexpat.errors import XML_ERROR_BAD_CHAR_REF 
import sys 
import math
import datetime
from turtle import bgcolor
from PyQt6 import QtCore, QtGui, QtWidgets
from PyQt6.QtGui import QPalette,QPixmap, QIcon
from PyQt6.QtWidgets import QMainWindow,QMessageBox,QFileDialog

from PyQt6 import QtSql
from PyQt6.QtSql import QSqlQuery

from openpyxl import load_workbook,Workbook 
from openpyxl.styles import Alignment,PatternFill,Color,Border,Font,NamedStyle,Side,borders
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.formatting.formatting import ConditionalFormattingList


import warnings
warnings.filterwarnings('ignore')

BASE_DIR= os.path.dirname(os.path.dirname(os.path.abspath(__file__) ) )
sys.path.append( BASE_DIR  )    
from ui.Ui_WireConvertForm import Ui_WireConvertForm 

class WireConvertForm(QMainWindow,Ui_WireConvertForm):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.setupUiEx()
        self.addConnect()
        
        self.db = QtSql.QSqlDatabase.addDatabase('QSQLITE')
        self.db.setDatabaseName(os.path.join(BASE_DIR,'db\\mdm.db'))

        self.initData()

    def closeEvent(self, QCloseEvent):   
        if self.db.isOpen:
            self.db.close()            

    def __wireaccSql(self,classtype):
        sql = 'select classtype, \
            spec, \
            itemno, \
            itemname, \
            unit \
            from wireacc \
            where classtype=\''       
        sql += classtype 
        sql += '\'' 
        return sql

    def __oricablelistSql(self,classtype):
        sql = 'select classtype, \
            spec, \
            itemno, \
            itemname, \
            voltagelevel, \
            voltagelevel2, \
            remark \
            from oricablelist \
            where classtype=\''       
        sql += classtype 
        sql += '\'' 
        return sql


    def setupUiEx(self):
        icon = QIcon()
        appPath=os.path.join(BASE_DIR,u'res\\icon\\WireConvert.ico')
        icon.addPixmap(QPixmap(appPath))
        self.setWindowIcon(icon)

    def addConnect(self):
        self.btnIFile.clicked.connect(self.btnIFileClick)
        self.btnOFile.clicked.connect(self.btnOFileClick)
        self.btnStart.clicked.connect(self.btnStartClick)

    def initData(self): 
        if not self.db.isOpen():
            if not self.db.open():               
                QMessageBox.critical(self, '线束表转化', self.db.lastError().text())
                return

    def btnIFileClick(self):
        fNames= QFileDialog.getOpenFileName(self,'线束表转化', '/','Excel File (*.xlsx)')
        if not fNames[0]:
            return
        if self.lineEOFile.text() == fNames[0]:
            QMessageBox.critical(self,'线束表转化', '输入与输出文件不能为同一文件') 
            self.lineEIFile.setFocus()
            return
        self.lineEIFile.setText(fNames[0])
    
    def btnOFileClick(self):
        fNames= QFileDialog.getSaveFileName(self,'线束表转化', '/','Excel File (*.xlsx)')
        if not fNames[0]:
            return
        
        if self.lineEIFile.text() == fNames[0]:
            QMessageBox.critical(self,'线束表转化', '输入与输出文件不能为同一文件') 
            self.lineEOFile.setFocus()
            return
        self.lineEOFile.setText(fNames[0])
    
    def btnStartClick(self):

        if len(self.lineEIFile.text())<=0 :
            QMessageBox.critical(self,'线束表转化', '输入文件为空') 
            self.lineEIFile.setFocus()
            return
        if len(self.lineEOFile.text())<=0:
            QMessageBox.critical(self,'线束表转化', '输入与输出文件不能为同一文件') 
            self.lineEOFile.setFocus()
            return
        if self.lineEIFile.text() == self.lineEOFile.text() :
            QMessageBox.critical(self,'线束表转化', '输入与输出文件不能为同一文件') 
            self.lineEOFile.setFocus()
            return
                

        self.convert(self.lineEIFile.text(),self.lineEOFile.text())

    def __getCoresFromProperty(self,s):
        if len(s) == 0:
            return ''
        if s.find("G") >= 0 :
            return s[0:s.find("G")]
        elif s.find("x") >= 0 :            
            return s[0:s.find("x")]
        return ''
    def __getSizeProperty(self,s):
        if len(s) == 0:
            return ''
        t = self.__rTrimUnit(s)
        if t.find("G") >= 0 :
            return t[t.find("G")+1:len(t)]
        elif s.find("x") >= 0 :            
            return t[t.find("x"):len(t)]
        return ''

    def __getRCode(self,s):
        if len(s) == 0:
            return ''
        if s.find("-") >= 0 :
            return s[s.find("-")+1:]
        return ''

    def __rTrimUnit(self,s):
        if len(s) == 0:
            return ''
        s = s.replace('mm²','')
        s = s.replace('mm','')
        if s.find(' ') >= 0 :
            s = s[0:s.find(' ')] 
        if len(s)>2:
            if s[len(s)-2:len(s)] =='.0':
                s = s[0:len(s)-2]
        return s.strip()         

    def isNumber(self,s):
        if len(s) == 0:
            return False
        if s.count(".")==1:  #小数的判断
            if s[0] == "-":
                s=s[1:]
            if s[0] == ".":
                return False
            s=s.replace(".","")
            for i in s:
                if i not in "0123456789":
                    return False
            return True
        elif s.count(".")==0:  #整数的判断
            if s[0]=="-":
                s=s[1:]
            for i in s:
                if i not in "0123456789":
                    return False
            return True
        else:
            return False 

    def fillstyle(self, sRow, eRow, sColumn, eColumn, ws):
        if eRow<sRow or eColumn<sColumn:
            return
        if eRow<1 or eColumn<1:
            return
        #左上
        ws.cell(sRow,sColumn).border = Border(left=Side(style='thick'),top=Side(style='thick'))
        #右上
        ws.cell(sRow,eColumn).border = Border(right=Side(style='thick'),top=Side(style='thick'))
        #左下
        ws.cell(eRow,sColumn).border = Border(left=Side(style='thick'),bottom=Side(style='thick'))
        #右下
        ws.cell(eRow,eColumn).border = Border(right=Side(style='thick'),bottom=Side(style='thick'))
        #上
        for i in range(sColumn+1,eColumn):
            ws.cell(sRow,i).border = Border(top=Side(style='thick'))
        #下
        for i in range(sColumn+1,eColumn):
            ws.cell(eRow,i).border = Border(bottom=Side(style='thick'))
        #左
        for i in range(sRow+1,eRow):
            ws.cell(i,sColumn).border = Border(left=Side(style='thick'))
        #右
        for i in range(sRow+1,eRow):
            ws.cell(i,eColumn).border = Border(right=Side(style='thick'))

    def convert(self,originalFile,targetFile):
         
        wib = load_workbook(filename=originalFile,read_only=True,data_only=True) 
        wb = Workbook()    
        startTime = datetime.datetime.now() 
        try:                                              
            wis=wib.active    
         
            iStarRow = self.spbStart.value()
            if iStarRow < 1:
                iStarRow = 1  
             
            iEndRow = self.spbEnd.value()                 
            if iEndRow ==0 or wis.max_row < iEndRow:
                iEndRow = wis.max_row   

            if iStarRow > iEndRow:
                QMessageBox.warning(self,'线束表转化', '选择的文件:' + originalFile + '及配置无需要转换的数据')
                wib.close()
                return   
            
            #检查关键字段不能为空            #
            self.progressBar.setMinimum(iStarRow)
            self.progressBar.setMaximum(iEndRow)
            self.lblProgress.setText(u'检查文件')
            iCount = 0
            strCabledt=''
            for iRow in range(iStarRow,iEndRow+1):
                self.progressBar.setValue(iRow)
                if wis['A' + str(iRow)].value is None:
                    QMessageBox.critical(self, '线束表转化', originalFile + '文件第【' + str(iRow) +'】行存在空数据或不正确数据')       
                    wib.close
                    wb.close
                    return
                if strCabledt == str(wis['A' + str(iRow)].value):   #电缆号
                    continue
                strCabledt = str(wis['A' + str(iRow)].value)
                if wis['A' + str(iRow)].value is None or \
                    wis['B' + str(iRow)].value is None or \
                    wis['D' + str(iRow)].value is None or \
                    wis['F' + str(iRow)].value is None or \
                    wis['G' + str(iRow)].value is None:        
                        QMessageBox.critical(self, '线束表转化', originalFile + '文件第【' + str(iRow) +'】行存在空数据或不正确数据')       
                        wib.close
                        wb.close
                        return       

            if not self.db.isOpen():
                if not self.db.open():               
                    QMessageBox.critical(self, '线束表转化', self.db.lastError().text())
                    wib.close
                    wb.close
                    return

            query = QSqlQuery() 
            ws = wb.active 
            ws.title = u'线束表转化' 
            
            ws = wb.create_sheet(u'线束辅料')     
            
            iRow = 1 
            ws.cell(iRow,1).value = '类别'  
            ws.cell(iRow,2).value = '规格'
            ws.cell(iRow,3).value = '物料号'   
            ws.cell(iRow,4).value = '名称'  
            ws.cell(iRow,5).value = '单位'
            iGXStart = 2
            #管型预绝缘端子            
            if not query.exec(self.__wireaccSql(u'管型预绝缘端子')):
                QMessageBox.critical(self,'线束表转化', query.lastError().text())
                wib.close
                wb.close
                return
            while query.next():    
                iRow += 1         
                ws.cell(iRow,1).value = str(query.value('classtype'))  
                ws.cell(iRow,2).value = str(query.value('spec'))   
                ws.cell(iRow,3).value = str(query.value('itemno'))   
                ws.cell(iRow,4).value = str(query.value('itemname'))  
                ws.cell(iRow,5).value = str(query.value('unit'))  
            
            #DT端子            
            if not query.exec(self.__wireaccSql(u'DT端子')):
                QMessageBox.critical(self,'线束表转化', query.lastError().text())
                wib.close
                wb.close
                return
            while query.next():    
                iRow += 1         
                ws.cell(iRow,1).value = str(query.value('classtype'))  
                ws.cell(iRow,2).value = str(query.value('spec'))   
                ws.cell(iRow,3).value = str(query.value('itemno'))   
                ws.cell(iRow,4).value = str(query.value('itemname'))  
                ws.cell(iRow,5).value = str(query.value('unit'))             
            #OT端子            
            if not query.exec(self.__wireaccSql(u'OT端子')):
                QMessageBox.critical(self,'线束表转化', query.lastError().text())
                wib.close
                wb.close
                return
            while query.next():    
                iRow += 1         
                ws.cell(iRow,1).value = str(query.value('classtype'))  
                ws.cell(iRow,2).value = str(query.value('spec'))   
                ws.cell(iRow,3).value = str(query.value('itemno'))   
                ws.cell(iRow,4).value = str(query.value('itemname'))  
                ws.cell(iRow,5).value = str(query.value('unit')) 
            iGXEnd = iRow
            strGXDataRange = '=线束辅料!$D$' + str(iGXStart) +':$D$' + str(iGXEnd)
            dvGX = DataValidation(type='list',formula1=strGXDataRange,allowBlank=True,prompt=u'选择绝缘端子')
            #热缩管
            if not query.exec(self.__wireaccSql(u'热缩管')):
                QMessageBox.critical(self,'线束表转化', query.lastError().text())
                return
            iRow = iGXEnd 
            iYSGStart = iGXEnd +1
            while query.next():   
                iRow += 1                    
                ws.cell(iRow,1).value = str(query.value('classtype'))  
                ws.cell(iRow,2).value = str(query.value('spec'))   
                ws.cell(iRow,3).value = str(query.value('itemno'))   
                ws.cell(iRow,4).value = str(query.value('itemname'))  
                ws.cell(iRow,5).value = str(query.value('unit'))  
            iYSGEnd = iRow                    
            strYSGataRange = '=线束辅料!$D$' + str(iYSGStart) +':$D$' + str(iYSGEnd)
            dvYSG = DataValidation(type='list',formula1=strYSGataRange,allowBlank=True,prompt=u'选择热缩管')
            #线标
            if not query.exec(self.__wireaccSql(u'线标')):
                QMessageBox.critical(self,'线束表转化', query.lastError().text())
                return
            iRow = iYSGEnd 
            iXBStart = iYSGEnd + 1
            while query.next():   
                iRow += 1                    
                ws.cell(iRow,1).value = str(query.value('classtype'))  
                ws.cell(iRow,2).value = str(query.value('spec'))   
                ws.cell(iRow,3).value = str(query.value('itemno'))   
                ws.cell(iRow,4).value = str(query.value('itemname'))  
                ws.cell(iRow,5).value = str(query.value('unit'))  
            iXBEnd = iRow                    
            strXBDataRange = '=线束辅料!$D$' + str(iXBStart) +':$D$' + str(iXBEnd)
            dvXB = DataValidation(type='list',formula1=strXBDataRange,allowBlank=True,prompt=u'选择线标')
            #扎带
            if not query.exec(self.__wireaccSql(u'扎带')):
                QMessageBox.critical(self,'线束表转化', query.lastError().text())
                return
            iRow = iXBEnd 
            iZDStart = iYSGEnd + 1
            while query.next():   
                iRow += 1                    
                ws.cell(iRow,1).value = str(query.value('classtype'))  
                ws.cell(iRow,2).value = str(query.value('spec'))   
                ws.cell(iRow,3).value = str(query.value('itemno'))   
                ws.cell(iRow,4).value = str(query.value('itemname'))  
                ws.cell(iRow,5).value = str(query.value('unit'))  
            iZDEnd = iRow                    
            strZDDataRange = '=线束辅料!$D$' + str(iZDStart) +':$D$' + str(iZDEnd)
            dvZD = DataValidation(type='list',formula1=strZDDataRange,allowBlank=True,prompt=u'选择扎带')
            
            #原缆清单
            ws = wb.create_sheet(u'原缆清单')      
            strType = self.cmbType.currentText()
            if not query.exec(self.__oricablelistSql(strType)):
                QMessageBox.critical(self,'线束表转化', query.lastError().text())
                wib.close
                wb.close
                return
            iRow = 1
            
            ws.cell(iRow,1).value = '类别'  
            ws.cell(iRow,2).value = '规格'
            ws.cell(iRow,3).value = '物料号'   
            ws.cell(iRow,4).value = '名称'  
            ws.cell(iRow,5).value = '电压' 
            ws.cell(iRow,6).value = '电压图纸' 
            ws.cell(iRow,7).value = '备注'
            while query.next():    
                iRow += 1         
                ws.cell(iRow,1).value = str(query.value('classtype'))  
                ws.cell(iRow,2).value = str(query.value('spec'))   
                ws.cell(iRow,3).value = str(query.value('itemno'))   
                ws.cell(iRow,4).value = str(query.value('itemname'))  
                ws.cell(iRow,5).value = str(query.value('voltagelevel'))  
                ws.cell(iRow,6).value = str(query.value('voltagelevel2'))  
                ws.cell(iRow,7).value = str(query.value('remark'))  

            #开始转换 
            ws = wb[u'线束表转化']
            ws.column_dimensions['A'].width=5
            ws.column_dimensions['B'].width=20
            ws.column_dimensions['C'].width=5
            ws.column_dimensions['D'].width=50
            ws.column_dimensions['E'].width=20
            ws.column_dimensions['F'].width=10 

            # 红色 #E72512
            # 黑色 #0A0A0D
            # 白色 #FFFFFF
            # 黄色 #F5FF00
            black = PatternFill(fill_type='solid',bgColor='0a0a0d') 
            white=PatternFill(fill_type='solid',bgColor='ffffff') 
            red=PatternFill(fill_type='solid',bgColor='e72512') 
            yellow=PatternFill(fill_type='solid',bgColor='f5ff00') 
            fontwhite= Font(name=u'等线',size=9,color='ffffff')
            fontblack=Font(name=u'等线',size=9,color='0a0a0d') 

            strCabledt = ''
            iCount = 0
            iYLQDDV = 1
            self.progressBar.setMinimum(iStarRow)
            self.progressBar.setMaximum(iEndRow)            
            self.lblProgress.setText(u'线束转化')
            for iRow in range(iStarRow,iEndRow+1):
                self.progressBar.setValue(iRow)
                if strCabledt == str(wis['A' + str(iRow)].value):   #电缆号
                    continue
                iCount += 1
                if wis['A' + str(iRow)].value is None:
                    strCabledt = ''
                else:
                    strCabledt = str(wis['A' + str(iRow)].value)                         
                if wis['B' + str(iRow)].value is None:        #芯线数与线径
                    strProperty =''
                else:
                    strProperty = wis['B' + str(iRow)].value
                if wis['D' + str(iRow)].value is None:        #外径
                    strLength =''
                else:
                    strLength = wis['D' + str(iRow)].value
                if wis['F' + str(iRow)].value is None:        #起始元件功能
                    strStarDevFun =''
                else:
                    strStarDevFun = wis['F' + str(iRow)].value                
                if wis['G' + str(iRow)].value is None:        #起始元件
                    strStarDevName =''
                else:
                    strStarDevName = wis['G' + str(iRow)].value
                  
                oCurRow = (iCount - 1) * 14 + 1          # 第一行  电缆组件

                ws.cell(oCurRow,1).value = '1'               #级别号
                #ws.cell(oCurRow,2).value = ''               #第二列、物料号
                ws.cell(oCurRow,3).value = 'L'               #第三列 
                ws.cell(oCurRow,4).value = u'电缆组件_' + strCabledt + '_' + strStarDevFun  #第四列
                ws.cell(oCurRow,5).value = '1'               #第五列 
                ws.cell(oCurRow,6).value = 'PC'               #第六列 
                ws.cell(oCurRow,7).value = ''                #第七列 
                 
                oCurRow += 1         
                oDupRow = oCurRow + 12                       # 第二行 and 第十四行   绝缘端子
                ws.cell(oCurRow,1).value = '2'               #级别号
                ws.cell(oCurRow,2).value = '=VLOOKUP(D' + str(oCurRow) +',IF({1,0},线束辅料!$D$' + str(iGXStart) +':$D$' + str(iGXEnd) +',线束辅料!$C$' + str(iGXStart) +':$C$' + str(iGXEnd) + '),2,FALSE)'               #第二列、无物料号
                ws.cell(oCurRow,3).value = 'L'               #第三列 
                dvGX.add(ws.cell(oCurRow,4))
                strTemp = self.__getSizeProperty(strProperty)
                if(strTemp == '0.34'):
                    ws.cell(oCurRow,4).value =  u'管形预绝缘端子_A0.34-10ET'  #第四列
                elif (strTemp=='0.5'):
                    ws.cell(oCurRow,4).value =  u'管形预绝缘端子_A0.5-10ET'  #第四列 
                elif(strTemp =='1'):
                    ws.cell(oCurRow,4).value =  u'带绝缘欧式端子_A1-10ET'  #第四列
                elif(strTemp =='1.5'):
                    ws.cell(oCurRow,4).value =  u'管形预绝缘端子_A1.5-10ET'  #第四列
                elif(strTemp =='2.5'):
                    ws.cell(oCurRow,4).value =  u'管形预绝缘端子_A2.5-10ET'  #第四列
                elif(strTemp =='16'):
                    ws.cell(oCurRow,4).value =  u'管形预绝缘端子_A16-10ET'  #第四列
                elif(strTemp =='25'):
                    ws.cell(oCurRow,4).value =  u'接线端子_KRF25-8'  #第四列
                elif(strTemp =='50'):
                    ws.cell(oCurRow,4).value =  u'接线端子_KRF50-10'  #第四列
                strTemp = self.__getCoresFromProperty(strProperty)
                ws.cell(oCurRow,5).value = strTemp               #第五列  
                ws.cell(oCurRow,6).value = '=VLOOKUP(D' + str(oCurRow) +',IF({1,0},线束辅料!$D$' + str(iGXStart) +':$D$' + str(iGXEnd) +',线束辅料!$E$' + str(iGXStart) +':$E$' + str(iGXEnd) + '),2,FALSE)'               #第六列 
                ws.cell(oCurRow,7).value = ' '                #第七列  
                for icolumn in range(1,8):
                    ws.cell(oDupRow,icolumn).value = '='  + get_column_letter(icolumn) + str(oCurRow)
                 
                oCurRow += 1                                # 第三行 and 第十三行   热缩管
                oDupRow = oCurRow + 10
                ws.cell(oCurRow,1).value = '2'               #级别号
                ws.cell(oCurRow,2).value = '=VLOOKUP(D' + str(oCurRow) +',IF({1,0},线束辅料!$D$' + str(iYSGStart) +':$D$' + str(iYSGEnd) +',线束辅料!$C$' + str(iYSGStart) +':$C$' + str(iYSGEnd) + '),2,FALSE)'               #第二列、无物料号
                ws.cell(oCurRow,3).value = 'L'               #第三列 
                dvYSG.add(ws.cell(oCurRow,4))
                strTemp = self.__rTrimUnit(strLength)
                if self.isNumber(strTemp): 
                    if float(strTemp) <= 9.5:     
                        ws.cell(oCurRow,4).value =  u'热缩管_D9,5/4,8'
                    elif float(strTemp) <= 12.7:     
                        ws.cell(oCurRow,4).value =  u'热缩管_D12.7/6.4mm_Black'
                    elif float(strTemp) <= 19.0:     
                        ws.cell(oCurRow,4).value =  u'热缩管_Φ19.0/9.5mm_Black'
                    elif float(strTemp) <= 30.0:     
                        ws.cell(oCurRow,4).value =  u'热缩管_Φ30/15mm_Black'
                    elif float(strTemp) <= 50.8:     
                        ws.cell(oCurRow,4).value =  u'热缩管_Ф50.8/25.4mm_Black'
                
                ws.cell(oCurRow,5).value = strLength               #第五列 
                ws.cell(oCurRow,6).value = '=VLOOKUP(D' + str(oCurRow) +',IF({1,0},线束辅料!$D$' + str(iYSGStart) +':$D$' + str(iYSGEnd) +',线束辅料!$E$' + str(iYSGStart) +':$E$' + str(iYSGEnd) + '),2,FALSE)'               #第六列 
                ws.cell(oCurRow,7).value = ' '                 #第七列 
                for icolumn in range(1,8):
                    ws.cell(oDupRow,icolumn).value = '='  + get_column_letter(icolumn) + str(oCurRow)
                
                oCurRow += 1                            # 第四行 and 第十行  线标 标签 
                oDupRow = oCurRow + 6
                ws.cell(oCurRow,1).value = '2'               #级别号
                ws.cell(oCurRow,2).value = '=VLOOKUP(D' + str(oCurRow) +',IF({1,0},线束辅料!$D$' + str(iXBStart) +':$D$' + str(iXBEnd) +',线束辅料!$C$' + str(iXBStart) +':$C$' + str(iXBEnd) + '),2,FALSE)'               #第二列、无物料号
                ws.cell(oCurRow,3).value = 'L'               #第三列 
                dvXB.add(ws.cell(oCurRow,4))
                ws.cell(oCurRow,4).value =  u'标签_40X27MM White'  #第四列
                ws.cell(oCurRow,5).value = '1'               #第五列 
                ws.cell(oCurRow,6).value = '=VLOOKUP(D' + str(oCurRow) +',IF({1,0},线束辅料!$D$' + str(iXBStart) +':$D$' + str(iXBEnd) +',线束辅料!$E$' + str(iXBStart) +':$E$' + str(iXBEnd) + '),2,FALSE)'               #第六列 
                ws.cell(oCurRow,7).value = ' '                 #第七列 
                for icolumn in range(1,8):
                    ws.cell(oDupRow,icolumn).value = '='   + get_column_letter(icolumn) + str(oCurRow)
                
                #第五、六， 十一，十二 为标签
                oCurRow += 1                            # 第五、六， 十一，十二 为标签
                oDupRow = oCurRow + 6                
                ws.merge_cells(start_row=oCurRow,start_column=4,end_row=oCurRow+1,end_column=4)
                ws.cell(oCurRow, 4).value = strCabledt + '-' + self.__getRCode(strStarDevName) + '\n' + strStarDevFun  #第四列                
                ws.cell(oCurRow, 4).alignment = Alignment(horizontal='center',vertical='center',wrapText=True)
                rule1  = FormulaRule(formula=['$B' + str(oCurRow -1) +'="DFB00052144"'],fill=black,stopIfTrue=True,font=fontwhite)
                rule2  = FormulaRule(formula=['$B' + str(oCurRow -1) +'="DFB00052145"'],fill=white,stopIfTrue=True,font=fontblack)
                rule3  = FormulaRule(formula=['$B' + str(oCurRow -1) +'="DFB00052146"'],fill=red,stopIfTrue=True,font=fontblack)
                rule4  = FormulaRule(formula=['$B' + str(oCurRow -1) +'<>""'],fill=yellow,stopIfTrue=True,font=fontblack)
                
                ws.conditional_formatting.add('$D$' + str(oCurRow),rule1)                
                ws.conditional_formatting.add('$D$' + str(oCurRow),rule2)
                ws.conditional_formatting.add('$D$' + str(oCurRow),rule3)
                ws.conditional_formatting.add('$D$' + str(oCurRow),rule4)

                ws.merge_cells(start_row=oDupRow,start_column=4,end_row=oDupRow+1,end_column=4)
                ws.cell(oDupRow, 4).value = '=D' + str(oCurRow) 
                #ws.cell(oDupRow, 4).value = strCabledt + '\n' + self.__getRCode(strStarDevName) + '\n' + strStarDevFun  #第四列   
                ws.cell(oDupRow, 4).alignment = Alignment(horizontal='center',vertical='center',wrapText=True)
                rule1  = FormulaRule(formula=['$B' + str(oDupRow -1) +'="DFB00052144"'],fill=black,stopIfTrue=True,font=fontwhite)
                rule2  = FormulaRule(formula=['$B' + str(oDupRow -1) +'="DFB00052145"'],fill=white,stopIfTrue=True,font=fontblack)
                rule3  = FormulaRule(formula=['$B' + str(oDupRow -1) +'="DFB00052146"'],fill=red,stopIfTrue=True,font=fontblack)
                rule4  = FormulaRule(formula=['$B' + str(oDupRow -1) +'<>""'],fill=yellow,stopIfTrue=True,font=fontblack)
                ws.conditional_formatting.add('$D$' + str(oDupRow),rule1)
                ws.conditional_formatting.add('$D$' + str(oDupRow),rule2)
                ws.conditional_formatting.add('$D$' + str(oDupRow),rule3)
                ws.conditional_formatting.add('$D$' + str(oDupRow),rule4)
                
                oCurRow += 2                            # #第七，第九  扎带 
                oDupRow = oCurRow + 2
                ws.cell(oCurRow,1).value = '2'               #级别号
                ws.cell(oCurRow,2).value = '=VLOOKUP(D' + str(oCurRow) +',IF({1,0},线束辅料!$D$' + str(iZDStart) +':$D$' + str(iZDEnd) +',线束辅料!$C$' + str(iZDStart) +':$C$' + str(iZDEnd) + '),2,FALSE)'               #第二列、无物料号
                ws.cell(oCurRow,3).value = 'L'               #第三列 
                dvZD.add(ws.cell(oCurRow,4))
                ws.cell(oCurRow,4).value = u'扎带L188φ48H1.3D4.8'       #第四列
                strTemp = self.__rTrimUnit(strLength)        
                if self.isNumber(strTemp): 
                    if float(strTemp) <= 25.0:     
                        ws.cell(oCurRow,4).value =  u'电缆扎带 Cable tie 2.5×100 MM'
                ws.cell(oCurRow,5).value = '1'               #第五列 
                ws.cell(oCurRow,6).value = '=VLOOKUP(D' + str(oCurRow) +',IF({1,0},线束辅料!$D$' + str(iZDStart) +':$D$' + str(iZDEnd) +',线束辅料!$E$' + str(iZDStart) +':$E$' + str(iZDEnd) + '),2,FALSE)'               #第六列 
                ws.cell(oCurRow,7).value = ' '                 #第七列 
                for icolumn in range(1,8):
                    ws.cell(oDupRow,icolumn).value = '='  + get_column_letter(icolumn) + str(oCurRow)
                
                oCurRow += 1            #第八行
                ws.cell(oCurRow,1).value = '2'               #级别号
                ws.cell(oCurRow,2).value = ''               #第二列、物料号
                ws.cell(oCurRow,3).value = 'L'               #第三列                 
                #ws.cell(oCurRow,4).value = '=VLOOKUP(B' + str(oCurRow) +',IF({1,0},原缆清单!$C:$C,原缆清单!$D:$D),2,FALSE)'               #第二列、物料号
                ws.cell(oCurRow,4).value = '=VLOOKUP(B' + str(oCurRow) +',原缆清单!$C:$D,2,FALSE)'               #第二列、物料号
                ws.cell(oCurRow,5).value = '1'               #第五列 
                ws.cell(oCurRow,6).value = 'M'               #第六列 
                ws.cell(oCurRow,7).value = strLength                #第七列 

                if len(strProperty)>0:
                    sql = self.__oricablelistSql(strType)
                    sql += ' and spec like \'' + self.__rTrimUnit(strProperty) + '%\''                
                    if not query.exec(sql):
                        QMessageBox.critical(self,'线束表转化', query.lastError().text())
                        wib.close
                        wb.close
                        return
                    iTemp = 0 
                    strFormula=''
                    while query.next():   
                        if len(strFormula) > 0:
                            strFormula +=','
                        else:
                            ws.cell(oCurRow,2).value = str(query.value('itemno'))
                        strFormula += '\'' 
                        strFormula += str(query.value('itemno'))

                    if strFormula.find(',')>0:
                        strFormula = '"' + strFormula + '"'
                        dvYL = DataValidation(type='list',formula1=strFormula,allowBlank=True,prompt=u'原缆清单')
                        dvYL.add(ws.cell(oCurRow,2))
                        ws.add_data_validation(dvYL)
                
                self.fillstyle((oCurRow-7),(oCurRow+6),1,7,ws)

            ws.add_data_validation(dvGX)
            ws.add_data_validation(dvYSG)
            ws.add_data_validation(dvXB)
            ws.add_data_validation(dvZD)

            wb.save(targetFile)
            wb.close
            wib.close
            
            endTime = datetime.datetime.now() 
            QMessageBox.information(self,'线束表转化','导出数据完成，文件名：' + targetFile +',总耗时：' + str(endTime-startTime))    
        except (NameError,ZeroDivisionError):
            wib.close
            wb.close
            QMessageBox.critical(self, '线束表转化', '变量名错误或除数为0')
        except OSError as reason:
            wib.close
            wb.close
            QMessageBox.critical(self, '线束表转化', str(reason))
        except TypeError as reason:
            wib.close
            wb.close
            QMessageBox.critical(self, '线束表转化', str(reason))
        except :
            wib.close
            wb.close
            QMessageBox.information(self,'线束表转化','导出数据文件失败')     