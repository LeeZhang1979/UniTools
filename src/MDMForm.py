#!/usr/bin/env python3
# coding=utf-8

'''MDMForm 系统配置主窗口'''

import os 
import sys 
from PyQt5 import QtCore, QtGui, QtWidgets 
from PyQt5.QtGui import QPalette, QPixmap, QIcon
from PyQt5.QtWidgets import QMainWindow,QMessageBox,QTableWidgetItem,QFileDialog

from PyQt5 import QtSql
from PyQt5.QtSql import QSqlQuery

from openpyxl import load_workbook,Workbook 

import warnings
warnings.filterwarnings('ignore')


BASE_DIR= os.path.dirname(os.path.dirname(os.path.abspath(__file__) ) )
sys.path.append( BASE_DIR  )   
from ui.Ui_MDMForm import Ui_MDMForm  

class MDMForm(QMainWindow,Ui_MDMForm):
    
    def __tablesql(self):
        """返回所有配置表的信息,用来对照Excel模板文件信息

        Returns:
            string: 查询表SQL语句
        """        
        sql= 'SELECT object_id, \
                object_name, \
                object_name_cn, \
                object_desc, \
                template_file, \
                template_sheet, \
                start_row, \
                end_row \
                FROM xt_objects \
                where object_type=\'T\' \
                order by object_id ASC'
        return sql

    def __columnsql(self,id):
        """指定表的列及对应excel内汉字名称

        Args:
            id (int): 配置的表ID

        Returns:
            string: SQL查询列信息
        """
        sql= 'SELECT \
            object_name, \
            object_name_cn, \
            object_desc, \
            column_mapping \
            FROM xt_objects \
            where object_type=\'C\' \
            and parent_object_id= '
        sql += str(id)
        
        #sql += ' and rim(column_mapping) !=\'\' '
        sql += ' order by column_mapping asc'

        return sql

    def __columns(self,id):
        """指定表的列及对应excel内汉字名称

        Args:
            id (int): 配置的表ID

        Returns:
            dict: 回指定表的列及对应excel内汉字名称字典
        """        
        cols = dict()
        query = QSqlQuery()                
        if not query.exec(self.__columnsql(str(id))):
            QMessageBox.critical(self,'MDM', query.lastError().text()) 
        else:   
            while query.next(): 
                cols[query.value('object_name')] = query.value('object_name_cn')
        return cols
    
    def __columnmap(self,id):
        """指定表的列及对应excel内对应列

        Args:
            id (int): 配置的表ID

        Returns:
            dict:  回指定表的列及对应excel内对应excel的列
        """        
        cols = dict()
        query = QSqlQuery()                
        if not query.exec(self.__columnsql(str(id))):
            QMessageBox.critical(self,'MDM', query.lastError().text()) 
        else:   
            while query.next(): 
                cols[query.value('object_name')] = query.value('column_mapping')
        return cols
    
    def __insertsql(self,columnmap,tbname):
        """插入配置表语句

        Args:
            columnmap (dict): 配置表字段与Excel对应列
            tbname (string): 配置表名

        Returns:
            string: 带参数的sql insert语句
        """        
        sqlk = ''          
        sqlv = ''
        for k,v in columnmap.items():
            if len(str(v))>0: 
                if len(sqlk) > 0:
                    sqlk +=','
                    sqlv +=','
                sqlk += str(k)     
                sqlv += '?'          
        sql = 'insert into '
        sql += tbname 
        sql += ' (' 
        sql += sqlk
        sql += ') values ('
        sql += sqlv    
        sql +=')'   
        return sql

    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.setupUiEx()
        self.addConnect()
        self.db = QtSql.QSqlDatabase.addDatabase('QSQLITE')
        self.db.setDatabaseName(os.path.join(BASE_DIR,'db\\mdm.db'))
        if not self.db.isOpen():
            if not self.db.open():               
                QMessageBox.critical(self, 'MDM', self.db.lastError().text())
                return
        self.initData()     
        if self.MDMListWidget.count()>0:
            self.MDMListWidget.setCurrentItem(self.MDMListWidget.item(0))
            self.mdmListClick()
            
    def closeEvent(self, QCloseEvent):   
        if self.db.isOpen:
            self.db.close()

    def setupUiEx(self):
        palette = QPalette()
        icon = QIcon()
        appPath=os.path.join(BASE_DIR,u'res\\imgs\\mdmconf.ico')
        icon.addPixmap(QPixmap(appPath))
        self.setWindowIcon(icon)

    def addConnect(self):
        self.MDMListWidget.clicked.connect(self.mdmListClick) 
        self.btnTemplate.clicked.connect(self.templateClick)
        self.btnImport.clicked.connect(self.importClick)
        self.btnExport.clicked.connect(self.exportClick)
        self.btnUpdate.clicked.connect(self.updateClick)
        
    def initData(self): 
        self.MDMListWidget.clear()
        query = QSqlQuery()
        if query.exec(self.__tablesql()):
            while query.next():       
                qItem = QtWidgets.QListWidgetItem()
                cols = dict()
                cols['object_id'] = query.value('object_id')
                cols['object_name'] = query.value('object_name')
                cols['object_name_cn'] = query.value('object_name_cn')
                cols['object_desc'] = query.value('object_desc')
                cols['template_file'] = query.value('template_file')
                cols['template_sheet'] = query.value('template_sheet')
                cols['start_row'] = query.value('start_row')
                cols['end_row'] = query.value('end_row')
                qItem.setData(1,cols)                
                qItem.setText(query.value('object_name_cn')) 
                
                self.MDMListWidget.addItem(qItem)      
        else:
             QMessageBox.critical(self,'MDM', query.lastError().text())           

    def showData(self,id,name):
        """绑定配置数据到界面

        Args:
            id (int): 配置表ID
            name (string): 配置表名
        """        
        cols = self.__columns(str(id))
        self.dataTableWidget.clear()
        self.dataTableWidget.setRowCount(0)
        self.dataTableWidget.setColumnCount(len(cols))
        self.dataTableWidget.setHorizontalHeaderLabels(cols.values())
        
        sql = ''
        for col in cols.keys():
            if(len(sql))>0:
                sql += ','
            sql += str(col)
        sql = 'select ' + sql
        sql +=' from '
        sql += str(name)        
        query = QSqlQuery()
        if not query.exec(sql):
            QMessageBox.critical(self,'MDM', query.lastError().text())
        else:
            while query.next():
                rows=self.dataTableWidget.rowCount()
                self.dataTableWidget.insertRow(rows)
                for i in range(len(cols)):
                    qtitem=QTableWidgetItem(str(query.value(list(cols.keys())[i])))
                    self.dataTableWidget.setItem(rows,i,qtitem)

    def mdmListClick(self):
        qItem=self.MDMListWidget.currentItem()          
        tconfs = dict(qItem.data(1))

        self.dataLabel.setText(str(tconfs['object_name_cn']) + ' : ' + str(tconfs['object_desc']))
        self.temFile.setText(str(tconfs['template_file']))
        self.temSheet.setText(str(tconfs['template_sheet']))
        self.temStart.setText(str(tconfs['start_row']))
        self.temEnd.setText(str(tconfs['end_row']))
        self.showData(str(tconfs['object_id']),str(tconfs['object_name']))

    def templateClick(self):
        if self.MDMListWidget.count()<=0:
            QMessageBox.information(self,'MDM', '请先选择要打开配置文件对应的基础数据配置表')
            return
        qItem=self.MDMListWidget.currentItem()
        cols = dict(qItem.data(1))

        if str(cols['template_file']) =='':
            QMessageBox.information(self,'MDM', '当前基础数据表尚未配置对应的配置文件')
            return
        appPath=os.path.join(BASE_DIR,str(cols['template_file']))
        #subprocess.run(appPath)
        os.system('start ' + appPath)
        #os.startfile(appPath) 
        return
    
    def importClick(self):
        if self.MDMListWidget.count()<=0:
            QMessageBox.information(self,'MDM', '请先选择要重新导入数据的基础数据配置表')
            return
        fNames= QFileDialog.getOpenFileName(self,'导入基础数据', '/','Excel File (*.xlsx)')
        if not fNames[0]:
            return
        
        qItem=self.MDMListWidget.currentItem()
        tconfs = dict(qItem.data(1))        
        sheetName = str(tconfs['template_sheet'])
        if QMessageBox.question(self, 'MDM', '确认更新模板配置表[' +sheetName + ']的数据?',QMessageBox.Yes|QMessageBox.No) == QMessageBox.No:
            return

        startRow = 2  #默认没有设置起始值，则默认从第二行开始（第一行为标题）
        if str(tconfs['start_row']).isdigit():
            startRow=int(str(tconfs['start_row']))                    
        endRow = 0   #没有设置结束行，默认后面数据行全部加载
        if str(tconfs['end_row']).isdigit():
            endRow=int(str(tconfs['end_row'])) 
        columnMap =dict()
        columnMap = self.__columnmap(str(tconfs['object_id']))
        
        try:  
            wb= load_workbook(fNames[0],True)                         
            if not (wb.sheetnames.index(sheetName) >= 0):
                QMessageBox.warning(self,'MDM', '选择的文件:' + fNames[0] + ',未包含配置指定的Sheet[' +sheetName + ']')
                wb.close()
                return        
            
            ws=wb[sheetName]        

            if endRow == 0:
                endRow = ws.max_row   # type: ignore  
            sql = 'delete from ' + str(tconfs['object_name'])                 
            query = QSqlQuery()
            if not query.exec_(sql):
                QMessageBox.warning(self,'MDM', '清空数据表[' + str(tconfs['object_name_cn'] + ':' + query.lastQuery() + ']失败' + query.lastError().text()))
                wb.close()
                return 
            sql = self.__insertsql(columnMap,str(tconfs['object_name']) )            
            query.prepare(sql)
            for iRow in (range(startRow,endRow+1)):   
                for k,v in columnMap.items():
                    if len(str(v))<=0:
                        continue
                    query.addBindValue(str(ws[str(v)+str(iRow)].value)) # type: ignore
                    
                if not query.exec():
                    QMessageBox.warning(self,'MDM', '执行语句[' + query.lastQuery() + ']失败,' + query.lastError().text())
                    wb.close()
                    return             
            wb.close()
            self.showData(str(tconfs['object_id']),str(tconfs['object_name']))
            QMessageBox.information(self,'MDM', '导入数据[' + str(tconfs['object_name_cn'])+ ']完成') 
        except: 
            QMessageBox.information(self,'MDM', '导入数据失败') 
          
    def exportClick(self):
        if self.MDMListWidget.count()<=0:
            QMessageBox.information(self,'MDM', '请先选择要重新导入数据的基础数据配置表')
            return
        fNames= QFileDialog.getSaveFileName(self,'下载基础数据', '/','Excel File (*.xlsx)')
        if not fNames[0]:
            return      
        qItem=self.MDMListWidget.currentItem()
        tconfs = dict(qItem.data(1))        
        sheetName = str(tconfs['template_sheet'])
        startRow = 2  #默认没有设置起始值，则默认从第二行开始（第一行为标题）
        if str(tconfs['start_row']).isdigit():
            startRow=int(str(tconfs['start_row']))                    

        columnMap = dict()
        column = dict()
        columnMap = self.__columnmap(str(tconfs['object_id']))
        column = self.__columns(str(tconfs['object_id']))

        try:            
            wb = Workbook()
            ws = wb.active
            ws.title = sheetName
            for k,v in columnMap.items():
                if len(str(v))<=0:
                    continue
                ws[str(v)+str(startRow-1)] = column[k]   # type: ignore 
            sql = ''
            for col in columnMap.keys():
                if(len(sql))>0:
                    sql += ','
                sql += str(col)
            sql = 'select ' + sql
            sql +=' from '
            sql += str(tconfs['object_name'])
            query = QSqlQuery()
            if not query.exec(sql):
                QMessageBox.critical(self,'MDM', query.lastError().text())
                return
            iRow = startRow
            while query.next():
                for k,v in columnMap.items():
                    if len(str(v))<=0:
                        continue
                    ws[str(v)+str(iRow)] = str(query.value(str(k))) # type: ignore 
                iRow += 1            
            wb.save(fNames[0])
            wb.close
            QMessageBox.information(self,'MDM','导出数据完成，文件名：' + fNames[0])    
            
        except: 
            QMessageBox.information(self,'MDM','导出数据文件失败，可能是文件类型错误') 
        return

    def updateClick(self):
        if self.MDMListWidget.count()<=0:
            QMessageBox.information(self,'MDM', '请先选择要更新模板文件数据的基础数据配置表')
            return
            
        qItem=self.MDMListWidget.currentItem()
        tconfs = dict(qItem.data(1))             
        if str(tconfs['template_file']) =='':
            QMessageBox.information(self,'MDM', '当前基础数据表尚未配置对应的配置文件')
            return
        tempfile=os.path.join(BASE_DIR,str(tconfs['template_file']))

        sheetName = str(tconfs['template_sheet'])   
        
        if QMessageBox.question(self, 'MDM', '确认更新本地模板文件:' + tempfile + ',配置表[' + sheetName + ']的数据?',QMessageBox.Yes|QMessageBox.No) == QMessageBox.No:
            return

        startRow = 2  #默认没有设置起始值，则默认从第二行开始（第一行为标题）
        if str(tconfs['start_row']).isdigit():
            startRow=int(str(tconfs['start_row']))                    

        try:               
            columnMap = dict()
            column = dict()
            columnMap = self.__columnmap(str(tconfs['object_id']))
            column = self.__columns(str(tconfs['object_id']))         
            wb = load_workbook(tempfile,False) 
            if not (wb.sheetnames.index(sheetName) >= 0):
                QMessageBox.warning(self,'MDM', '选择的文件:' + tempfile + ',未包含配置指定的Sheet[' + sheetName + ']')
                wb.close()
                return        
            ws=wb[sheetName]     
            #maxRow = ws.max_row # type: ignore 
            # #暂未实现清除文档中老数据（考虑有附加列未导入数据库，如图片等）
            '''
            if startRow >1: 
                for k,v in columnMap.items():
                    if len(str(v))<=0:
                        continue
                    QMessageBox.information(self,'MDM', str(column[k]))
                    ws[str(v)+str(startRow-1)] = str(column[k])   # type: ignore 
                    #更新标题暂未实现
            '''
            sql = ''
            for col in columnMap.keys():
                if(len(sql))>0:
                    sql += ','
                sql += str(col)
            sql = 'select ' + sql
            sql +=' from '
            sql += str(tconfs['object_name'])
            query = QSqlQuery()
            if not query.exec(sql):
                QMessageBox.critical(self,'MDM', query.lastError().text())
                return
            iRow = startRow
            while query.next():
                for k,v in columnMap.items():
                    if len(str(v))<=0:
                        continue 
                    ws[str(v)+str(iRow)] = str(query.value(str(k))) # type: ignore 
                iRow += 1            
            wb.save(tempfile)
            wb.close
            QMessageBox.information(self,'MDM','更新模板文件数据:' + tempfile + '完成')    
            
        except: 
            QMessageBox.information(self,'MDM','更新模板文件数据失败，可能是文件类型错误') 
        return