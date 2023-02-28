#!/usr/bin/env python3
# coding=utf-8

'''动力电缆计算器'''

import os 
import sys 
import math
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtGui import QPalette,QPixmap, QIcon, QImage, QRegExpValidator,QColor
from PyQt5.QtWidgets import QMainWindow,QMessageBox,QTableWidgetItem,QFileDialog
from PyQt5.QtCore import QRegExp 

from PyQt5 import QtSql
from PyQt5.QtSql import QSqlQuery

from openpyxl import Workbook 
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

import warnings
warnings.filterwarnings('ignore')

BASE_DIR= os.path.dirname(os.path.dirname(os.path.abspath(__file__) ) )
sys.path.append( BASE_DIR  )    
from ui.Ui_PowerCableForm import Ui_PowerCableForm 

class PowerCableForm(QMainWindow,Ui_PowerCableForm):
    
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.setupUiEx()
        self.addConnect()         
        
        self.db = QtSql.QSqlDatabase.addDatabase('QSQLITE')
        self.db.setDatabaseName(os.path.join(BASE_DIR,'db\\mdm.db'))

        self.initData()

    def closeEvent(self, QCloseEvent):   
        if self.db.isOpen:
            self.db.close()            

    def setupUiEx(self):
        icon = QIcon()
        appPath=os.path.join(BASE_DIR,u'res\\icon\\powercable.ico')
        icon.addPixmap(QPixmap(appPath))
        self.setWindowIcon(icon)

        self.cmbCrossType.setIconSize(QtCore.QSize(53, 64))
        self.cmbCrossType.setItemIcon(0, QIcon(os.path.join(BASE_DIR,u'res\\imgs\\multi_2core.png')))
        self.cmbCrossType.setItemIcon(1, QIcon(os.path.join(BASE_DIR,u'res\\imgs\\multi_3core.png')))
        self.cmbCrossType.setItemIcon(2, QIcon(os.path.join(BASE_DIR,u'res\\imgs\\single_2core_t.png')))
        self.cmbCrossType.setItemIcon(3, QIcon(os.path.join(BASE_DIR,u'res\\imgs\\single_3core_t.png')))
        self.cmbCrossType.setItemIcon(4, QIcon(os.path.join(BASE_DIR,u'res\\imgs\\signle_3core_ft.png')))
        self.cmbCrossType.setItemIcon(5, QIcon(os.path.join(BASE_DIR,u'res\\imgs\\single_3core_hs.png')))
        self.cmbCrossType.setItemIcon(6, QIcon(os.path.join(BASE_DIR,u'res\\imgs\\single_3core_vs.png')))
        self.refreshLayingPic()

    def refreshLayingPic(self):
        filename=''
        if self.cmbLayingType.currentIndex() == 0:
            if self.rbTouch.isChecked():
                filename=os.path.join(BASE_DIR,u'res\\imgs\\tray_t_h.png') 
            elif self.rbSpace.isChecked():
                filename=os.path.join(BASE_DIR,u'res\\imgs\\tray_s_h.png') 
        elif self.cmbLayingType.currentIndex() == 1:
            if self.rbTouch.isChecked():
                filename=os.path.join(BASE_DIR,u'res\\imgs\\tray_t_v.png') 
            elif self.rbSpace.isChecked():
                filename=os.path.join(BASE_DIR,u'res\\imgs\\tray_s_v.png') 
        elif self.cmbLayingType.currentIndex() == 2:
            if self.rbTouch.isChecked():
                filename=os.path.join(BASE_DIR,u'res\\imgs\\lorc_t.png') 
            elif self.rbSpace.isChecked():
                filename=os.path.join(BASE_DIR,u'res\\imgs\\lorc_s.png')  
        qimage = QImage(filename)
        self.lblLayingImg.setPixmap(QPixmap.fromImage(qimage))                
        self.lblLayingImg.setScaledContents(True)

    def addConnect(self):
        #单绕组电流(A) 5位以内>0整数
        intReg = QRegExp('^[1-9][0-9]{1,4}')
        regExpValidator = QRegExpValidator(intReg)
        self.lineEEC.setValidator(regExpValidator)
        #绕组数 3位以内>0整数
        intReg = QRegExp('^[1-9][0-9]{1,2}')
        regExpValidator = QRegExpValidator(intReg) 
        self.lineEWings.setValidator(regExpValidator) 
        #长度 5位以内>=0浮点数 +2位小数
        floatReg = QRegExp('^([0]|[1-9][0-9]{0,4})(?:\.\d{1,2})?$|(^\t?$)')
        regExpValidator = QRegExpValidator(floatReg)
        self.lineELong.setValidator(regExpValidator) 
        #单价 5位以内>=0浮点数 +2位小数
        floatReg = QRegExp('^([0]|[1-9][0-9]{0,4})(?:\.\d{1,2})?$|(^\t?$)')
        regExpValidator = QRegExpValidator(floatReg)
        self.lineEPrice.setValidator(regExpValidator) 
        #标称截面积(㎜²) 5位以内>=0浮点数 +2位小数
        floatReg = QRegExp('^([0]|[1-9][0-9]{0,4})(?:\.\d{1,2})?$|(^\t?$)')
        regExpValidator = QRegExpValidator(floatReg)
        self.lineECS.setValidator(regExpValidator) 
        #环温(°C)  4位以内>=0整数
        intReg = QRegExp('^([0]|[1-9][0-9]{1,3})')
        regExpValidator = QRegExpValidator(intReg)
        self.lineEAmbientT.setValidator(regExpValidator)  
        #护套耐温(°C) 4位以内>=0整数
        intReg = QRegExp('^([0]|[1-9][0-9]{1,3})')
        regExpValidator = QRegExpValidator(intReg)  
        self.lineESTR.setValidator(regExpValidator)   
        #托盘/梯架数 1,2,3
        intReg = QRegExp('[1-3]?')
        regExpValidator = QRegExpValidator(intReg)  
        self.lineENumber.setValidator(regExpValidator)  
        #三相回路数 1,2,3
        intReg = QRegExp('[1-3]?')
        regExpValidator = QRegExpValidator(intReg)  
        self.lineECircuits.setValidator(regExpValidator)   

        self.rbTouch.clicked.connect(self.refreshLayingPic)
        self.rbSpace.clicked.connect(self.refreshLayingPic)
        self.cmbLayingType.currentIndexChanged.connect(self.refreshLayingPic)
        self.btnCalculation.clicked.connect(self.btnCalculationClick)
        self.btnAdd.clicked.connect(self.btnAddClick)
        self.btnUpdate.clicked.connect(self.btnUpdateClick)
        self.btnDelete.clicked.connect(self.btnDeleteClick)
        self.btnExport.clicked.connect(self.btnExportClick)
        self.tblList.itemDoubleClicked.connect(self.tblSelected)
    
    def initData(self):  
        self.cleanResult()        
        self.tblList.clearContents()
        self.tblList.setRowCount(0)
    
    def isNumber(self,s):
        if len(s) == 0:
            return False
        if s.count(".")==1:  #小数的判断
            if s[0] == "-":
                s=s[1:]
            if s[0] == ".":
                return False
            s=s.replace(".","")
            for i in s:
                if i not in "0123456789":
                    return False
            return True
        elif s.count(".")==0:  #整数的判断
            if s[0]=="-":
                s=s[1:]
            for i in s:
                if i not in "0123456789":
                    return False
            return True
        else:
            return False

    def __ampacitySQL(self, conductor,material,sheathtr,crosssection):
        sql= 'select conductor, \
            material, \
            sheath_tr, \
            ambientt, \
            cross_section, \
            multi_2core, \
            multi_3core, \
            single_2core_t, \
            single_3core_t, \
            signle_3core_ft, \
            single_3core_hs, \
            single_3core_vs \
            from ampacity \
            where ambientt=\'30\' and \
            conductor = \''
        sql += conductor
        sql += '\' and material = \''
        sql += material 
        sql += '\' and sheath_tr = \''        
        sql += sheathtr 
        sql += '\' and cross_section = \''        
        sql += crosssection 
        sql += '\'' 
        return sql
    
    def __ambienttcfSQL(self, insulatedType,ambientt):
        sql= 'select insulated_type, \
            ambientt, \
            pvc, \
            xlpe_epr, \
            mineral_ipvct, \
            mineral_ic \
            from ambienttcf \
            where insulated_type = \''
        sql += insulatedType
        sql += '\' and ambientt = \''
        sql += ambientt 
        sql += '\'' 
        return sql

    def __layingcfSQL(self,layingtype,touchtype,layingcount):
        sql= 'select laying_type, \
            touch_type, \
            laying_count, \
            circuits_laying1, \
            circuits_laying2, \
            circuits_laying3, \
            remark \
            from layingcf \
            where laying_type = \''
        sql += layingtype
        sql += '\' and touch_type = \''
        sql += touchtype 
        sql += '\' and laying_count = \''        
        sql += layingcount 
        sql += '\'' 
        return sql

    def cleanResult(self):
        #载流量值
        self.lblOECNum.setText('')
        #敷设系数
        self.lblOACF.setText('')
        #折算系数
        self.lblOCFnum.setText('')
        #单绕组电缆根数
        self.lblOSNum.setText('')
        #向上取整
        self.lblOSNumUp.setText('')   
        #单相电缆根数
        self.lblOSDNum.setText('')     
        #单绕组电流余量
        self.lblOECLeftNum.setText('')
        #余量百分比
        self.lblOECPer.setText('')
        #电缆成本
        self.lblOCost.setText('')
        #当前最低成本
        self.lblOLowestCost.setText('')  

    def calculation(self):
        self.cleanResult()
        strPara = ''
        #电缆类型
        self.cmbPCType.currentText()
        #单绕组电流(A)
        if self.isNumber(self.lineEEC.text()):
            paraX = float(self.lineEEC.text())
        else:
            QMessageBox.critical(self,'动力电缆计算', '单绕组电流为空或非数字') 
            self.lineEEC.setFocus()
            return False
        
        #长度
        linelong = float(0.0)
        if self.isNumber(self.lineELong.text()):
            linelong = float(self.lineELong.text())
        #单价
        price = float(0.0)
        if self.isNumber(self.lineEPrice.text()):
            price = float(self.lineEPrice.text())      

        #绕组数
        if self.isNumber(self.lineEWings.text()):
            paraC = float(self.lineEWings.text())
        else:
            QMessageBox.critical(self,'动力电缆计算', '绕组数为空或非数字') 
            self.lineEWings.setFocus()
            return False
        #Query ampacity Table
        #导体
        conductor = self.cmbConductor.currentText()
        #标称截面积(㎜²)
        crosssection = self.lineECS.text()
        #绝缘材料
        material = self.cmbMaterial.currentText()
        #护套耐温(°C)
        sheathtr = self.lineESTR.text()
        #电缆芯数及排列
        #self.cmbCrossType.currentText()
        paraY = float(1.0)
        if not self.db.isOpen():
            if not self.db.open():               
                QMessageBox.critical(self, '动力电缆计算', self.db.lastError().text())
                return
        query = QSqlQuery()
        if query.exec(self.__ampacitySQL(conductor,material,sheathtr,crosssection)):
            while query.next():
                if self.cmbCrossType.currentIndex() == 0:
                    strPara = str(query.value('multi_2core'))
                elif self.cmbCrossType.currentIndex() == 1:
                    strPara = str(query.value('multi_3core'))
                elif self.cmbCrossType.currentIndex() == 2:
                    strPara = str(query.value('single_2core_t'))
                elif self.cmbCrossType.currentIndex() == 3:
                    strPara = str(query.value('single_3core_t'))
                elif self.cmbCrossType.currentIndex() == 4:
                    strPara = str(query.value('signle_3core_ft'))
                elif self.cmbCrossType.currentIndex() == 5:
                    strPara = str(query.value('single_3core_hs')) 
                elif self.cmbCrossType.currentIndex() == 6:
                    strPara = str(query.value('single_3core_vs'))
        else:
            QMessageBox.critical(self,'动力电缆计算', query.lastError().text())  
            return False
        if self.isNumber(strPara):
            paraY = float(strPara)        
        elif strPara == '\\':
            paraY = 1.0
        else:
            QMessageBox.critical(self,'动力电缆计算', '根据给出的条件,未找到对应额定载流量')  
            return False
        #Query ambienttcf Table
        insulatedType = u'绝缘'
        #环温(°C)
        ambientt = self.lineEAmbientT.text()
        strPara = ''
        paraA = float(0.0)
        if self.cmbCrossType.currentIndex() == 5:
            strPara = '1.0'  #硬性指定为1.0
        elif self.cmbCrossType.currentIndex() == 6:
            strPara = '1.0'  #硬性指定为1.0
        elif query.exec(self.__ambienttcfSQL(insulatedType,ambientt)):
            while query.next():
                if self.cmbMaterial.currentIndex() == 0:
                    strPara = str(query.value('pvc'))
                elif self.cmbMaterial.currentIndex() == 1:
                    strPara = str(query.value('xlpe_epr'))
                elif self.cmbMaterial.currentIndex() == 2:
                    strPara = str(query.value('mineral_ipvct'))
                elif self.cmbMaterial.currentIndex() == 3:
                    strPara = str(query.value('mineral_ic'))
        else:
            QMessageBox.critical(self,'动力电缆计算', query.lastError().text())  
            return False
        
        if self.isNumber(strPara):
            paraA = float(strPara)
        elif strPara == '\\':
            paraA = float(1.0)
        else:
            QMessageBox.critical(self,'动力电缆计算', '根据给出的条件,未找到对应折算系数')  
            return False
        #Query layingcf Table
        #敷设方式
        layingtype = self.cmbLayingType.currentText()
        #相互接触
        touchtype = ''
        if self.rbTouch.isChecked():
            touchtype = self.rbTouch.text()
        elif self.rbSpace.isChecked():
            #有间距
            touchtype = self.rbSpace.text()
         
        #托盘/梯架数
        layingcount = self.lineENumber.text()
        
        #三相回路数
        circuitslaying = self.lineECircuits.text() 
        strPara = ''
        paraB = float(0.0)
        if query.exec(self.__layingcfSQL(layingtype,touchtype,layingcount)):
            while query.next(): 
                if circuitslaying == '1':
                    strPara = str(query.value('circuits_laying1'))
                elif circuitslaying == '2':
                    strPara = str(query.value('circuits_laying2'))
                elif circuitslaying == '3':
                    strPara = str(query.value('circuits_laying3'))
        else:
            QMessageBox.critical(self,'动力电缆计算', query.lastError().text())  
            return False
        if self.isNumber(strPara):
            paraB = float(strPara)
        elif strPara == '\\':
            paraB = 1.0
        else:
            QMessageBox.critical(self,'动力电缆计算', '根据给出的条件,未找到对应敷设系数')  
            return False 

        temp = float(1.0)
        temp *= paraY  #载流量
        temp *= paraA  #温度折算系数
        temp *= paraB  #敷设系数
        #单绕组电缆根数
        paraZ = round(paraX / temp,2)
        #向上取整
        paraD = math.ceil(paraZ)
        #单相电缆根数
        paraF = paraD * paraC  #paraC 绕组数
        #单绕组电流余量
        paraE = round(paraY * paraD - paraX,2) 
        #余量百分比
        paraP = round(paraE / paraX,4)
        
        #载流量值
        self.lblOECNum.setText(str(paraY))
        #温度折算系数
        self.lblOCFnum.setText(str(paraA))
        #敷设系数
        self.lblOACF.setText(str(paraB))
        #单绕组电缆根数
        self.lblOSNum.setText(str(paraZ))
        #向上取整
        self.lblOSNumUp.setText(str(paraD))
        #单相电缆根数
        self.lblOSDNum.setText(str(paraF))
        #单绕组电流余量
        self.lblOECLeftNum.setText(str(paraE))
        #余量百分比
        self.lblOECPer.setText(f'{round(paraP*100,2)}%')
        #电缆成本
        self.lblOCost.setText(str(round(float(3.0)*paraF*price*linelong,2) ))
 
        return True

    def lowestPrice(self):
        if self.tblList.rowCount()<=0:
            self.lblOLowestCost.setText('')
            return
        lowestRow = 0
        result = float(0.0) 
        row = 0
        while row < self.tblList.rowCount():
            price = self.tblList.item(row,23).text()
            if self.isNumber(price):
                if row==0 or result > float(price):
                    result = float(price)       
                    lowestRow = row             
            row +=1

        price = self.tblList.item(lowestRow,23).text()
        result = float(price) 
        self.lblOLowestCost.setText(str(round(result,2)))
        self.lblO10.setText(u'最低成本第' + str(lowestRow+1) + u'行:')
        return 

    def btnCalculationClick(self):
        self.calculation()

    def btnAddClick(self):
        if not self.calculation():
            return
        #列表 
        newRow = self.tblList.rowCount()
        self.tblList.insertRow(newRow)

        #电缆类型
        self.tblList.setItem(newRow,0, QTableWidgetItem(self.cmbPCType.currentText()))
        #单绕组电流(A) 
        self.tblList.setItem(newRow,1, QTableWidgetItem(self.lineEEC.text()))
        #绕组数
        self.tblList.setItem(newRow,2, QTableWidgetItem(self.lineEWings.text()))
        #长度
        self.tblList.setItem(newRow,3, QTableWidgetItem(self.lineELong.text()))
        #单价
        self.tblList.setItem(newRow,4, QTableWidgetItem(self.lineEPrice.text()))
        #导体
        self.tblList.setItem(newRow,5, QTableWidgetItem(self.cmbConductor.currentText()))
        #标称截面积(㎜²)
        self.tblList.setItem(newRow,6, QTableWidgetItem(self.lineECS.text()))
        #绝缘材料 
        self.tblList.setItem(newRow,7, QTableWidgetItem(self.cmbConductor.currentText()))
        #环温(°C)  
        self.tblList.setItem(newRow,8, QTableWidgetItem(self.lineEAmbientT.text()))
        #护套耐温(°C) 
        self.tblList.setItem(newRow,9, QTableWidgetItem(self.lineESTR.text()))  
        #电缆芯数排列
        self.tblList.setItem(newRow,10, QTableWidgetItem(self.cmbCrossType.currentText()))
        #敷设方式
        self.tblList.setItem(newRow,11, QTableWidgetItem(self.cmbLayingType.currentText()))
        #是否接触
        touchtype=''
        if self.rbTouch.isChecked():
            touchtype = self.rbTouch.text()
        elif self.rbSpace.isChecked():
            #有间距
            touchtype = self.rbSpace.text()
        self.tblList.setItem(newRow,12, QTableWidgetItem(touchtype))
        #托盘/梯架数 
        self.tblList.setItem(newRow,13, QTableWidgetItem(self.lineENumber.text()))
        #三相回路数 
        self.tblList.setItem(newRow,14, QTableWidgetItem(self.lineECircuits.text()))
        #载流量值
        self.tblList.setItem(newRow,15, QTableWidgetItem(self.lblOECNum.text()))
        #折算系数
        self.tblList.setItem(newRow,16, QTableWidgetItem(self.lblOCFnum.text()))
        #敷设系数
        self.tblList.setItem(newRow,17, QTableWidgetItem(self.lblOACF.text()))
        #单绕组电缆根数
        self.tblList.setItem(newRow,18, QTableWidgetItem(self.lblOSNum.text()))
        #向上取整
        self.tblList.setItem(newRow,19, QTableWidgetItem(self.lblOSNumUp.text()))  
        #单相电缆根数
        self.tblList.setItem(newRow,20, QTableWidgetItem(self.lblOSDNum.text()))   
        #单绕组电流余量
        self.tblList.setItem(newRow,21, QTableWidgetItem(self.lblOECLeftNum.text()))
        #余量百分比
        self.tblList.setItem(newRow,22, QTableWidgetItem(self.lblOECPer.text()))
        #电缆成本
        self.tblList.setItem(newRow,23, QTableWidgetItem(self.lblOCost.text()))
         
        self.lowestPrice()
        #qtitem = QTableWidgetItem(self.cmbPCType.currentText())
        #self.cmbPCType.setCurrentText(text)
        #qtitem.setData(QtCore.Qt.ItemDataRole.UserRole, '') 

    def btnUpdateClick(self):
        row = self.tblList.currentRow()
        if row< 0:
            QMessageBox.critical(self,'动力电缆计算', '请双击要更新的行,并更新数据后，再点击更新选中的行') 
            return
        
        if not self.calculation():
            return
      
        #电缆类型 
        self.tblList.item(row,0).setText(self.cmbPCType.currentText())
        #单绕组电流(A) 
        self.tblList.item(row,1).setText(self.lineEEC.text())
        #绕组数
        self.tblList.item(row,2).setText(self.lineEWings.text())        
        #长度
        self.tblList.item(row,3).setText(self.lineELong.text())
        #单价
        self.tblList.item(row,4).setText(self.lineEPrice.text())
        #导体
        self.tblList.item(row,5).setText(self.cmbConductor.currentText())
        #标称截面积(㎜²)
        self.tblList.item(row,6).setText(self.lineECS.text())
        #绝缘材料 
        self.tblList.item(row,7).setText(self.cmbConductor.currentText())
        #环温(°C)  
        self.tblList.item(row,8).setText(self.lineEAmbientT.text())
        #护套耐温(°C) 
        self.tblList.item(row,9).setText(self.lineESTR.text())
        #电缆芯数排列
        self.tblList.item(row,10).setText(self.cmbCrossType.currentText())
        #敷设方式
        self.tblList.item(row,11).setText(self.cmbLayingType.currentText())
        #是否接触
        touchtype=''
        if self.rbTouch.isChecked():
            touchtype = self.rbTouch.text()
        elif self.rbSpace.isChecked():
            #有间距
            touchtype = self.rbSpace.text()
        self.tblList.item(row,12).setText(touchtype)
        #托盘/梯架数 
        self.tblList.item(row,13).setText(self.lineENumber.text())
        #三相回路数 
        self.tblList.item(row,14).setText(self.lineECircuits.text())
        #载流量值
        self.tblList.item(row,15).setText(self.lblOECNum.text())
        #折算系数
        self.tblList.item(row,16).setText(self.lblOCFnum.text())
        #敷设系数
        self.tblList.item(row,17).setText(self.lblOACF.text())
        #单绕组电缆根数
        self.tblList.item(row,18).setText(self.lblOSNum.text())
        #向上取整
        self.tblList.item(row,19).setText(self.lblOSNumUp.text())
        #单相电缆根数
        self.tblList.item(row,20).setText(self.lblOSDNum.text())
        #单绕组电流余量
        self.tblList.item(row,21).setText(self.lblOECLeftNum.text())
        #余量百分比
        self.tblList.item(row,22).setText(self.lblOECPer.text())    
        #电缆成本
        self.tblList.item(row,23).setText(self.lblOCost.text())

        self.lowestPrice()

    def btnDeleteClick(self):
        row = self.tblList.currentRow()
        if row<= 0:
            QMessageBox.critical(self,'动力电缆计算', '请选择要更新的行后，再点击删除选中的行') 
            return
        
        if QMessageBox.critical(self,'动力电缆计算', '确认删除当前行',QMessageBox.Yes|QMessageBox.No) == QMessageBox.No:
            return
        self.tblList.removeRow(row)

    def btnExportClick(self): 
        if self.tblList.rowCount()<= 0:
            QMessageBox.critical(self,'动力电缆计算', '请先添加需要导出的信息') 
            return
        fNames= QFileDialog.getSaveFileName(self,'生成动力电缆计算文件', '/','Excel File (*.xlsx)')
        if not fNames[0]:
            return
        try:            
            wb = Workbook() 
            ws = wb.active 
            ws.title = u'动力电缆计算'
            listRow = 0  
            while listRow < self.tblList.rowCount():
                row = 1
                column = listRow * 4 +1 
                ws.column_dimensions[get_column_letter(column)].width = 5
                ws.column_dimensions[get_column_letter(column+1)].width = 20
                ws.column_dimensions[get_column_letter(column+2)].width = 50
                 #电缆类型 
                ws.merge_cells(start_row=row,start_column=column,end_row=row,end_column=column+2)
                ws.cell(row, column).value = self.tblList.item(listRow,0).text()
                ws.cell(row, column).alignment = Alignment(horizontal='center',vertical='center')
                ws.merge_cells(start_row=row+1,start_column=column,end_row=row+11,end_column=column)
                ws.cell(row+1, column).value = u'输入'                   
                ws.cell(row+1, column).alignment = Alignment(horizontal='center',vertical='center')
                ws.merge_cells(start_row=row+12,start_column=column,end_row=row+20,end_column=column)
                ws.cell(row+12, column).value = u'输出'
                ws.cell(row+12, column).alignment = Alignment(horizontal='center',vertical='center')
                i = 1
                while i < self.tblList.columnCount(): 
                    ws.cell(row+i, column+1).value = self.tblList.horizontalHeaderItem(i).text() 
                    ws.cell(row+i, column+2).value = self.tblList.item(listRow,i).text()
                    i += 1
                listRow += 1
            wb.save(fNames[0])
            wb.close
            QMessageBox.information(self,'动力电缆计算','导出数据完成，文件名：' + fNames[0])    
        except (NameError,ZeroDivisionError):
            QMessageBox.critical(self, '动力电缆计算', '变量名错误或除数为0')
        except OSError as reason:
            QMessageBox.critical(self, '动力电缆计算', str(reason))
        except TypeError as reason:
            QMessageBox.critical(self, '动力电缆计算', str(reason))
        except :
            QMessageBox.information(self,'动力电缆计算','导出数据文件失败') 

        self.lowestPrice()
        return        
        
    def tblSelected(self):
        
        row = self.tblList.currentRow()
        if row<= 0:
            return
        #电缆类型
        self.cmbPCType.setCurrentText(self.tblList.item(row,0).text())
        #单绕组电流(A) 
        self.lineEEC.setText(self.tblList.item(row,1).text())
        #绕组数
        self.lineEWings.setText(self.tblList.item(row,2).text())
        #长度
        self.lineELong.setText(self.tblList.item(row,3).text())
        #单价
        self.lineEPrice.setText(self.tblList.item(row,4).text())
        #导体
        self.cmbConductor.setCurrentText(self.tblList.item(row,5).text())
        #标称截面积(㎜²)
        self.lineECS.setText(self.tblList.item(row,6).text())
        #绝缘材料 
        self.cmbConductor.setCurrentText(self.tblList.item(row,7).text())
        #环温(°C)  
        self.lineEAmbientT.setText(self.tblList.item(row,8).text())
        #护套耐温(°C) 
        self.lineESTR.setText(self.tblList.item(row,9).text())
        #电缆芯数排列
        self.cmbCrossType.setCurrentText(self.tblList.item(row,10).text())
        #敷设方式
        self.cmbLayingType.setCurrentText(self.tblList.item(row,11).text())
        #是否接触
        if self.tblList.item(row,12).text() == self.rbTouch.text() :
            self.rbTouch.setChecked(True)
        elif self.tblList.item(row,12).text() == self.rbSpace.text() :
            #有间距
            self.rbSpace.setChecked(True) 
        #托盘/梯架数 
        self.lineENumber.setText(self.tblList.item(row,13).text())
        #三相回路数 
        self.lineECircuits.setText(self.tblList.item(row,14).text())
        #载流量值
        self.lblOECNum.setText(self.tblList.item(row,15).text())
        #折算系数
        self.lblOCFnum.setText(self.tblList.item(row,16).text())
        #敷设系数
        self.lblOACF.setText(self.tblList.item(row,17).text())
        #单绕组电缆根数
        self.lblOSNum.setText(self.tblList.item(row,18).text())
        #向上取整
        self.lblOSNumUp.setText(self.tblList.item(row,19).text())  
        #单相电缆根数
        self.lblOSDNum.setText(self.tblList.item(row,20).text())
        #单绕组电流余量
        self.lblOECLeftNum.setText(self.tblList.item(row,21).text())
        #余量百分比
        self.lblOECPer.setText(self.tblList.item(row,22).text())
        #成本
        self.lblOCost.setText(self.tblList.item(row,23).text())
        